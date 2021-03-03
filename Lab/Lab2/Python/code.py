import openpyxl
import datetime
import json
import os.path
from flask import Flask, request
def writeexel():
    global line, buffer
    book = openpyxl.load_workbook('data.xlsx')
    sheet= book.active
    for items in buffer:
        for i in range(5):
            sheet.cell(line,i+1).value = items[i]
        line+=1
    sheet.cell(1,6).value = line-1
    book.save('data.xlsx')
    book.close
    buffer.clear()
app = Flask(__name__)  
@app.route('/', methods = ['POST', 'GET'])
def index():
    if request.method == 'POST':
        time = datetime.datetime.now().time()
        json_from_post= request.get_json()
        global number, buffer
        for item in json_from_post['cart']:
            buy = [number, json_from_post['user_id'], time, item['item'], item['price']]
            number += 1  
            buffer.append(buy)
        if len(buffer)>1000:
            writeexel()
        return 'OK'
if __name__ == "__main__":
    global number, buffer, line
    number = 1
    line = 2
    buffer = []
    if not(os.path.exists('data.xlsx')):    
        book = openpyxl.Workbook()
        sheet = book.active
        sheet.cell(1,1).value = 'N'
        sheet.cell(1,2).value = 'User ID'
        sheet.cell(1,3).value = 'Datetime'
        sheet.cell(1,4).value = 'Item'
        sheet.cell(1,5).value = 'Price'
        sheet.cell(1,6).value = 1
        book.save('data.xlsx')
        book.close
    else:
        book = openpyxl.load_workbook('data.xlsx')
        sheet = book.active
        number = sheet.cell(1,6).value
        book.close
    app.run()








#book = openpyxl.open('data.xlsx', read_only=True)
#    sheet = book.active
#    lastadd = sheet["Y25"].value
#    nest = sheet["Z25"].value
#    book.close
#    timereq = datetime.datetime.now().time()
#    if request.method == 'POST':
#        userid = request.json["user_id"]
#        cart = request.json["cart"]
#        book = openpyxl.load_workbook('data.xlsx')
#        sheet = book.active
#        lastadd+=1
#        sheet.cell(lastadd,columnN).value = nest
#        sheet.cell(lastadd,columnID).value = userid
#        sheet.cell(lastadd,columnTime).value = timereq
#        print(lastadd)
#        for item in cart:
#            itemname = item['item']
#            itemprice = item['price']
#            sheet.cell(lastadd,columnItem).value = itemname
#            sheet.cell(lastadd,columnPrice).value = itemprice
#            lastadd+=1
#        sheet.cell(25,25).value = lastadd-1
#        sheet.cell(25,26).value = nest+1
#        book.save('data.xlsx')
#        book.close
#        return 'OK'